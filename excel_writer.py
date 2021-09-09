from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, Alignment

class ExcelWriter:
    @staticmethod
    def generate_sheet(current_sheet, teams, provided_by):
        num_cols = 2 # very important variable, play with it

        thick = Side(border_style="thick")
        thin = Side(border_style="thin")

        for table, team in enumerate(teams):
            sci = 5 * (table % num_cols) + 2 # starting column index
            sri = (table // num_cols) * (len(team.matches) * 2 + 4) + 2 # starting row index

            current_sheet.merge_cells(f"{get_column_letter(sci)}{sri}:{get_column_letter(sci + 3)}{sri}")
            current_sheet[f"{get_column_letter(sci)}{sri}"].value = "Provided by {}".format(provided_by if provided_by != None else "Quantum Robotics")
            current_sheet[f"{get_column_letter(sci)}{sri}"].border = Border(top=thick, left=thick, right=thick, bottom=Side(border_style="thin", color="ffffff"))
            current_sheet[f"{get_column_letter(sci)}{sri}"].font = Font(sz=14)
            current_sheet[f"{get_column_letter(sci)}{sri}"].alignment = Alignment(horizontal="center", vertical="center")

            current_sheet.row_dimensions[3].height = 30
            current_sheet.merge_cells(f"{get_column_letter(sci)}{sri + 1}:{get_column_letter(sci + 3)}{sri + 1}")
            current_sheet[f"{get_column_letter(sci)}{sri + 1}"].value = "Team #" + str(team.number) + " | " + team.name
            current_sheet[f"{get_column_letter(sci)}{sri + 1}"].border = Border(top=Side(border_style="thin", color="ffffff"), left=thick, right=thick, bottom=thick)
            current_sheet[f"{get_column_letter(sci)}{sri + 1}"].font = Font(sz=24)
            current_sheet[f"{get_column_letter(sci)}{sri + 1}"].alignment = Alignment(horizontal="center", vertical="center")

            for column, column_name in enumerate(["Match", "Partner", "Opponent 1", "Opponent 2"]):
                current_sheet.column_dimensions[get_column_letter(sci + column)].width = 27

                current_sheet[f"{get_column_letter(sci + column)}{sri + 2}"].value = column_name
                current_sheet[f"{get_column_letter(sci + column)}{sri + 2}"].border = Border(top=thick, left=thin, right=thin, bottom=thin)
                current_sheet[f"{get_column_letter(sci + column)}{sri + 2}"].font = Font(sz=14)
                current_sheet[f"{get_column_letter(sci + column)}{sri + 2}"].alignment = Alignment(horizontal="center", vertical="center")

            for index, match in enumerate(team.matches):
                is_red = team in match["red_alliance"]

                current_cell = current_sheet[f"{get_column_letter(sci)}{sri + 2 * index + 3}"]
                current_cell.value = match["name"]
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=Side(border_style="thin", color="ffffff"))
                current_cell.font  = Font(sz=14, color="ff0000" if is_red else "0000ff")
                current_cell.alignment = Alignment(horizontal="center", vertical="center")
                current_cell = current_sheet[f"{get_column_letter(sci)}{sri + 2 * index + 4}"]
                current_cell.value = "Red" if is_red else "Blue"
                current_cell.border = Border(top=Side(border_style="thin", color="ffffff"), left=thin, right=thin, bottom=thin)
                current_cell.font  = Font(sz=14, color="ff0000" if is_red else "0000ff")
                current_cell.alignment = Alignment(horizontal="center", vertical="center")

                partner = next(member for member in (match["red_alliance"] if is_red else match["blue_alliance"]) if team != member)
                current_cell = current_sheet[f"{get_column_letter(sci + 1)}{sri + 2 * index + 3}"]
                current_cell.value = str(partner.number)
                current_cell.border = Border(top=thin, left=thin, right=thin, bottom=Side(border_style="thin", color="ffffff"))
                current_cell.font  = Font(sz=14, color="ff0000" if is_red else "0000ff")
                current_cell.alignment = Alignment(horizontal="center", vertical="center")
                current_cell = current_sheet[f"{get_column_letter(sci + 1)}{sri + 2 * index + 4}"]
                current_cell.value = partner.name
                current_cell.border = Border(top=Side(border_style="thin", color="ffffff"), left=thin, right=thin, bottom=thin)
                current_cell.font  = Font(sz=14, color="ff0000" if is_red else "0000ff")
                current_cell.alignment = Alignment(horizontal="center", vertical="center")

                for column, opponent in enumerate(match["blue_alliance"] if is_red else match["red_alliance"]): # +2
                    current_cell = current_sheet[f"{get_column_letter(sci + column + 2)}{sri + 2 * index + 3}"]
                    current_cell.value = str(opponent.number)
                    current_cell.border = Border(top=thin, left=thin, right=thin, bottom=Side(border_style="thin", color="ffffff"))
                    current_cell.font  = Font(sz=14, color="0000ff" if is_red else "ff0000")
                    current_cell.alignment = Alignment(horizontal="center", vertical="center")
                    current_cell = current_sheet[f"{get_column_letter(sci + column + 2)}{sri + 2 * index + 4}"]
                    current_cell.value = opponent.name
                    current_cell.border = Border(top=Side(border_style="thin", color="ffffff"), left=thin, right=thin, bottom=thin)
                    current_cell.font  = Font(sz=14, color="0000ff" if is_red else "ff0000")
                    current_cell.alignment = Alignment(horizontal="center", vertical="center")