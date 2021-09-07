from openpyxl import load_workbook

from team import Team

class ExcelLoader:
    def __init__(self, teams_file, matches_file):
        self.input_teams = load_workbook(filename=teams_file)["Sheet1"]
        self.input_matches = load_workbook(filename=matches_file)["Sheet1"]

    def load_teams(self):
        teams = []
        for team_number, team_name in zip(self.input_teams["A"], self.input_teams["B"]):
            teams.append(Team(int(team_number.value), team_name.value))

        return teams

    def load_matches(self, teams):
        matches = []
        for match_name, red_1, red_2, blue_1, blue_2 in zip(
            self.input_matches["A"], 
            self.input_matches["B"], 
            self.input_matches["C"], 
            self.input_matches["D"], 
            self.input_matches["E"]
        ):
            matches.append({
                "name": match_name.value, 
                "red_alliance": [next(team for team in teams if team.number == int(red_1.value)), next(team for team in teams if team.number == int(red_2.value))],
                "blue_alliance": [next(team for team in teams if team.number == int(blue_1.value)), next(team for team in teams if team.number == int(blue_2.value))]
            })

        return matches